from dash import Dash, dcc, html, dash_table, no_update, callback_context
import plotly.graph_objs as go
from dash.dependencies import State, Input, Output
from dash.exceptions import PreventUpdate
from openpyxl import load_workbook

import pandas as pd
import numpy as np
import plotly.express as px
import json

# File paths
file_path = 'data/Australian Shark-Incident Database Public Version.xlsx'
geojson_path = "data/australian-states.json"

# Efficiently load the Excel file using openpyxl
wb = load_workbook(filename=file_path, read_only=True)
sheet = wb.active

# Load all rows into a list
df_chunks = []
for row in sheet.iter_rows(values_only=True):
    df_chunks.append(row)

# Convert rows into a DataFrame while preserving the original structure
df_shark = pd.DataFrame(df_chunks[1:], columns=df_chunks[0])

# Drop rows with completely null values to avoid issues in plotting
df_shark.dropna(how='all', inplace=True)

# Optionally, clean specific columns (e.g., 'State') if they have null values
df_shark = df_shark[df_shark['State'].notnull()]

# Load GeoJSON file for Australian states
with open(geojson_path) as f:
    aus_states_geojson = json.load(f)

# Initialize the Dash app
app = Dash(
    __name__,
    meta_tags=[
        {
            "name": "viewport",
            "content": "width=device-width, initial-scale=1, maximum-scale=1.0, user-scalable=no",
        }
    ],
)
app.title = "Australian Shark Incident Analysis"
server = app.server
app.config["suppress_callback_exceptions"] = True

# Build the upper-left panel
def build_upper_left_panel():
    return html.Div(
        id="upper-left",
        className="three columns",
        children=[
            html.P(
                className="section-title",
                children="Choose filters to see shark incident data",
            ),
            html.Div(
                className="filter-section",
                children=[
                    html.Label(className="filter-section-title", children="Compare States"),
                    dcc.Dropdown(
                        id="state-select",
                        options=[{"label": i, "value": i} for i in df_shark['State'].unique()],
                        value=[df_shark['State'].unique()[0]],
                        multi=True,
                        placeholder="Select states to compare",
                        className="dropdown",
                    ),
                ],
            ),
            html.Div(
                className="filter-section",
                children=[
                    html.Label(className="filter-section-title", children="Choose Metrics to Compare"),
                    dcc.Dropdown(
                        id="metric-select",
                        options=[
                            {"label": "Victim Age", "value": "Victim.age"},
                            {"label": "Shark Length (m)", "value": "Shark.length.m"},
                            {"label": "Number of Sharks", "value": "No.sharks"},
                            {"label": "Fatal Incidents", "value": "fatal_count"},
                            {"label": "Total Incidents", "value": "incident_count"},
                        ],
                        value=["incident_count"],
                        multi=True,
                        className="dropdown",
                    ),
                ],
            ),
            html.Div(
                className="filter-section",
                children=[
                    html.Label(className="filter-section-title", children="Filter by Region"),
                    dcc.Checklist(
                        id="region-select-all",
                        options=[{"label": "Select All Regions", "value": "All"}],
                        value=[],
                        className="checkbox",
                    ),
                    dcc.Dropdown(
                        id="region-select",
                        multi=True,
                        searchable=True,
                        placeholder="Select specific regions",
                        className="dropdown",
                    ),
                ],
            ),
        ],
    )

# Create a parallel coordinates plot
def create_parallel_coordinates():
    numerical_cols = ['Victim.age', 'Shark.length.m', 'Latitude', 'Longitude', 'Incident.year']
    dimensions = []

    for col in numerical_cols:
        numeric_series = pd.to_numeric(df_shark[col], errors='coerce')
        dimensions.append(
            dict(
                range=[numeric_series.min(), numeric_series.max()],
                label=col.replace('.', ' '),
                values=numeric_series
            )
        )
    
    
    # Add categorical dimension for Shark Species

    # Add categorical dimension for Shark Species
    dimensions.append(
        dict(
            range=[0, len(df_shark['Shark.common.name'].unique())],
            ticktext=df_shark['Shark.common.name'].unique(),
            tickvals=list(range(len(df_shark['Shark.common.name'].unique()))),
            label='Shark Species',
            values=[list(df_shark['Shark.common.name'].unique()).index(x) for x in df_shark['Shark.common.name']]
        )
    )
    
    
    # Create parallel coordinates plot

    # Create parallel coordinates plot
    fig = go.Figure(data=
        go.Parcoords(
            line=dict(
                color=df_shark['Incident.year'],
                colorscale='Viridis'
            ),
            dimensions=dimensions,
        )
    )
    
    
    # Update layout

    # Update layout
    fig.update_layout(
        plot_bgcolor='#171b26',
        paper_bgcolor='#171b26',
        font=dict(color='#737a8d'),
        margin=dict(l=80, r=80, t=30, b=30),
    )
    
    return fig

# Callback to update the parallel coordinates plot dynamically
@app.callback(
    Output("parallel-coords-plot", "figure"),
    [
        Input("state-select", "value"),
        Input("year-slider", "value")
    ]
)
def update_parallel_coordinates(selected_states, year_range):
    if not selected_states:
        selected_states = df_shark['State'].unique()
        
    filtered_df = df_shark[
        (df_shark['State'].isin(selected_states)) &
        (df_shark['Incident.year'] >= year_range[0]) &
        (df_shark['Incident.year'] <= year_range[1])
    ]
    
    numerical_cols = ['Victim.age', 'Shark.length.m', 'Latitude', 'Longitude', 'Incident.year']
    dimensions = []
    
    
    # Create dimensions for numerical columns

    # Create dimensions for numerical columns
    for col in numerical_cols:
        numeric_series = pd.to_numeric(filtered_df[col], errors='coerce')
        dimensions.append(
            dict(
                range=[numeric_series.min(), numeric_series.max()],
                label=col.replace('.', ' '),
                values=numeric_series
            )
        )
    
    dimensions.append(
        dict(
            range=[0, len(filtered_df['State'].unique())],
            ticktext=filtered_df['State'].unique(),
            tickvals=list(range(len(filtered_df['State'].unique()))),
            label='State',
            values=[list(filtered_df['State'].unique()).index(x) for x in filtered_df['State']]
        )
    )
    
    fig = go.Figure(data=
        go.Parcoords(
            line=dict(
                color=filtered_df['Incident.year'],
                colorscale='Viridis'
            ),
            dimensions=dimensions
        )
    )
    fig.update_layout(
        title_text=f"Parallel Coordinates View ({year_range[0]}-{year_range[1]})",
        plot_bgcolor='#171b26',
        paper_bgcolor='#171b26',
        font=dict(color='#737a8d')
    )
    
    return fig

# Define the layout
app.layout = html.Div(
    className="container scalable",
    children=[
        html.Div(
            id="banner",
            className="banner",
            children=[
                html.H6("Australian Shark Attack"),
                html.Img(src=app.get_asset_url("plotly_logo_white.png")),
            ],
        ),
        html.Div(
            [
                # Main Content (Slider + Map)
                html.Div(
                    className="map-and-slider",
                    children=[
                        # Slider on top of the map
                        html.Div(
                            className="filter-section",
                            children=[
                                html.Label(className="filter-section-title", children="Year Range"),
                                dcc.RangeSlider(
                                    id="year-slider",
                                    min=df_shark['Incident.year'].min(),
                                    max=df_shark['Incident.year'].max(),
                                    step=1,
                                    marks={
                                        str(year): str(year)
                                        for year in range(
                                            int(df_shark['Incident.year'].min()),
                                            int(df_shark['Incident.year'].max()) + 1,
                                            10,
                                        )
                                    },
                                    value=[df_shark['Incident.year'].min(), df_shark['Incident.year'].max()],
                                    className="range-slider",
                                ),
                            ],
                        ),
                    ],
                ),
            ]
        ),
        html.Div(
            id="upper-container",
            className="row",
            children=[
                build_upper_left_panel(),
                html.Div(
                    id="geo-map-outer",
                    className="eight columns",
                    children=[
                        html.P(
                            id="map-title",
                            children="Shark Incidents in Australia",
                        ),
                        html.Div(
                            id="geo-map-loading-outer",
                            children=[
                                dcc.Loading(
                                    id="loading",
                                    children=dcc.Graph(
                                        id="geo-map",
                                        figure={
                                            "data": [],
                                            "layout": dict(
                                                plot_bgcolor="#171b26",
                                                paper_bgcolor="#171b26",
                                            ),
                                        },
                                    ),
                                )
                            ],
                        ),
                    ],
                ),
            ],
        ),
        html.Div(
            id="lower-container",
            className="row",
            children=[
                # Parallel Coordinates Plot
                html.Div(
                    className="six columns",
                    children=[
                        html.P(
                            className="section-title",
                            children="Parallel Coordinates View of Shark Incident Data",
                        ),
                        dcc.Graph(
                            id="parallel-coords-plot",
                            figure=create_parallel_coordinates()
                        ),
                    ],
                ),
                # Radar Plot
                html.Div(
                    className="six columns",
                    children=[
                        html.P(
                            className="section-title",
                            children="Radar Plot: Shark Incidents by Month and State",
                        ),
                        dcc.Graph(
                            id="radar-plot",
                            figure={}  # Initial empty figure
                        ),
                    ],
                ),
                html.Div(
                    id="splom-container",
                    children=[
                        html.P(
                            className="section-title",
                            children="Scatterplot Matrix (SPLOM) of Shark Incidents",
                        ),
                        dcc.Graph(
                            id="splom-chart",
                            figure={}  # Initial empty figure
                        ),
                    ],
                ),
                
            ],
        ),
    ],
)

# @app.callback(
#     Output("radar-plot", "figure"),
#     [Input("state-select", "value")]
# )
# def update_radar_plot(selected_states):
#     if not selected_states:
#         selected_states = df_shark["State"].unique()

#     # Filter data for selected states
#     filtered_data = df_shark[df_shark["State"].isin(selected_states)]

#     # Group data by month and state, and count the incidents
#     radar_data = (
#         filtered_data.groupby(['Incident.month', 'State'])
#         .size()
#         .reset_index(name='Incidents')  # 'Incidents' column now contains the count
#     )

#     # Create a pivot table to organize data for plotting
#     radar_pivot = radar_data.pivot(index='Incident.month', columns='State', values='Incidents').fillna(0)
#     radar_pivot = radar_pivot.reindex(range(1, 13), fill_value=0)  # Ensure all months are represented
#     radar_pivot.reset_index(inplace=True)
#     radar_pivot.rename(columns={'Incident.month': 'Month'}, inplace=True)

#     # Month mapping for consistent display
#     month_mapping = {
#         1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
#         5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
#         9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
#     }
#     radar_pivot['Month'] = radar_pivot['Month'].map(month_mapping)

#     # Melt the data for plotting
#     melted_data = radar_pivot.melt(id_vars='Month', var_name='State', value_name='Incidents')

#     # Debugging: Print the final melted data
#     print("Melted Data:", melted_data)

#     # Create the radar plot
#     fig = px.line_polar(
#         melted_data,
#         r='Incidents',
#         theta='Month',
#         color='State',
#         line_close=True,
#         title="Shark Incidents by Month and State"
#     )

#     # Adjust layout
#     fig.update_layout(
#         polar=dict(
#             angularaxis=dict(
#                 categoryorder='array',
#                 categoryarray=["Jan", "Feb", "Mar", "Apr", "May", "Jun",
#                                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]  # Explicit order
#             )
#         ),
#         legend=dict(
#             title="State",
#             orientation="h",
#             yanchor="bottom",
#             y=-0.2,
#             xanchor="center",
#             x=0.5
#         ),
#     )
#     return fig

# Callback for SPLOM chart
@app.callback(
    Output("splom-chart", "figure"),
    [Input("state-select", "value")]
)
def update_splom_chart(selected_states):
    # Filter data based on selected states
    if not selected_states:
        selected_states = df_shark["State"].unique()
    df_filtered = df_shark[df_shark["State"].isin(selected_states)]
    variables = ['Shark.common.name', 'Victim.activity', 'Provocative.act', 'Victim.injury']

    # Clean the data
    df_cleaned = df_filtered[variables]
    df_cleaned = df_cleaned[~df_cleaned['Victim.injury'].str.lower().eq('unknown')]
    df_cleaned['Victim.injury'] = df_cleaned['Victim.injury'].str.lower().replace(['injury', 'injured'], 'injured')
    df_cleaned['Victim.activity'] = df_cleaned['Victim.activity'].str.replace(r'^other:.*', 'other', case=False, regex=True)
    df_cleaned['Provocative.act'] = df_cleaned['Provocative.act'].replace(
        'victim intentionally moved into immediate proximity of shark',
        'move close to shark'
    )

    # Create the SPLOM chart
    fig = px.scatter_matrix(
        df_cleaned,
        dimensions=variables,
        title="Scatterplot Matrix (SPLOM) of Shark Incidents",
        labels={
            'Shark.common.name': 'Shark Name',
            'Victim.injury': 'Victim Injury',
            'Victim.activity': 'Victim Activity',
            'Provocative.act': 'Provocative Act'
        },
        color='Victim.injury',
    )
    fig.update_layout(
        dragmode='select',
        width=900,
        height=900,
        showlegend=True,
    )
    return fig


@app.callback(
    [Output("region-select", "options"), Output("region-select", "value")],
    [Input("state-select", "value"), Input("region-select-all", "value")],
)
def update_region_dropdown(state_select, select_all):
    if not state_select:
        return [], []
        
    # Filter data for the selected state(s)
    state_data = df_shark[df_shark["State"].isin([state_select] if isinstance(state_select, str) else state_select)]

    # Get unique regions and handle NaN values
    regions = state_data["Location"].fillna("Unknown Location").unique()
    
    # Convert all values to strings and filter out empty strings
    regions = [str(region) for region in regions if str(region).strip()]
    
    # Sort the cleaned list
    regions.sort()

    # Create dropdown options
    options = [{"label": region, "value": region} for region in regions]

    # Auto-select all regions if "Select All Regions" is checked
    if select_all and "All" in select_all:
        value = regions
    else:
        value = []

    return options, value


@app.callback(
    Output("checklist-container", "children"),
    [Input("region-select", "value")],
    [State("region-select", "options"), State("region-select-all", "value")],
)
def update_checklist(selected, select_options, checked):
    if len(selected) < len(select_options) and len(checked) == 0:
        raise PreventUpdate()

    elif len(selected) < len(select_options) and len(checked) == 1:
        return dcc.Checklist(
            id="region-select-all",
            options=[{"label": "Select All Regions", "value": "All"}],
            value=[],
        )

    elif len(selected) == len(select_options) and len(checked) == 1:
        raise PreventUpdate()

    return dcc.Checklist(
        id="region-select-all",
        options=[{"label": "Select All Regions", "value": "All"}],
        value=["All"],
    )

@app.callback(
    Output("cost-stats-container", "children"),
    [
        Input("geo-map", "selectedData"),
        Input("procedure-plot", "selectedData"),
        Input("metric-select", "value"),
        Input("state-select", "value"),
    ],
)
def update_hospital_datatable(geo_select, procedure_select, cost_select, state_select):
    state_agg = df_shark[df_shark['State'] == state_select]
    # make table from geo-select
    geo_data_dict = {
        "Location": [],
        "Shark Species": [],
        "Incident Year": [],
        "Maximum Metric": [],
        "Minimum Metric": [],
    }

    ctx = callback_context
    if ctx.triggered:
        prop_id = ctx.triggered[0]["prop_id"].split(".")[0]

        # make table from procedure-select
        if prop_id == "procedure-plot" and procedure_select is not None:
            for point in procedure_select["points"]:
                location = point["customdata"]
                dff = state_agg[state_agg["Location"] == location]

                if not dff.empty:
                    geo_data_dict["Location"].append(location)
                    shark_species = dff["Shark.common.name"].tolist()[0]
                    geo_data_dict["Shark Species"].append(shark_species)

                    year = dff["Incident.year"].tolist()[0]
                    geo_data_dict["Incident Year"].append(year)

                    geo_data_dict["Maximum Metric"].append(
                        dff[cost_select].max()
                    )
                    geo_data_dict["Minimum Metric"].append(
                        dff[cost_select].min()
                    )

        if prop_id == "geo-map" and geo_select is not None:
            for point in geo_select["points"]:
                location = point["customdata"][0]
                dff = state_agg[state_agg["Location"] == location]

                if not dff.empty:
                    geo_data_dict["Location"].append(location)
                    geo_data_dict["Shark Species"].append(dff["Shark.common.name"].tolist()[0])

                    year = dff["Incident.year"].tolist()[0]
                    geo_data_dict["Incident Year"].append(year)

                    geo_data_dict["Maximum Metric"].append(
                        dff[cost_select].max()
                    )
                    geo_data_dict["Minimum Metric"].append(
                        dff[cost_select].min()
                    )

        geo_data_df = pd.DataFrame(data=geo_data_dict)
        data = geo_data_df.to_dict("records")

    else:
        data = [{}]

    return dash_table.DataTable(
        id="cost-stats-table",
        columns=[{"name": i, "id": i} for i in geo_data_dict.keys()],
        data=data,
        filter_action="native",
        page_size=5,
        style_cell={"background-color": "#242a3b", "color": "#7b7d8d"},
        style_as_list_view=False,
        style_header={"background-color": "#1f2536", "padding": "0px 5px"},
    )

@app.callback(
    Output("procedure-stats-container", "children"),
    [
        Input("procedure-plot", "selectedData"),
        Input("geo-map", "selectedData"),
        Input("metric-select", "value"),
    ],
    [State("state-select", "value")],
)
def update_procedure_stats(procedure_select, geo_select, cost_select, state_select):
    procedure_dict = {
        "Shark Species": [],
        "Location": [],
        "Incident Year": [],
        "Metric Summary": [],
    }

    ctx = callback_context
    prop_id = ""
    if ctx.triggered:
        prop_id = ctx.triggered[0]["prop_id"].split(".")[0]

    if prop_id == "procedure-plot" and procedure_select is not None:
        for point in procedure_select["points"]:
            procedure_dict["Shark Species"].append(point["y"])
            procedure_dict["Location"].append(point["customdata"])
            procedure_dict["Incident Year"].append(point["x"])
            procedure_dict["Metric Summary"].append(("${:,.2f}".format(point["x"])))

    # Display all procedures at selected location
    location_select = []

    if prop_id == "geo-map" and geo_select is not None:
        for point in geo_select["points"]:
            location = point["customdata"][0]
            location_select.append(location)

        state_raw_data = df_shark[df_shark['State'] == state_select]
        location_filtered = state_raw_data[
            state_raw_data["Location"].isin(location_select)
        ]

        for i in range(len(location_filtered)):
            procedure_dict["Shark Species"].append(
                location_filtered.iloc[i]["Shark.common.name"]
            )
            procedure_dict["Location"].append(
                location_filtered.iloc[i]["Location"]
            )
            procedure_dict["Incident Year"].append(
                location_filtered.iloc[i]["Incident.year"]
            )
            procedure_dict["Metric Summary"].append(
                "${:,.2f}".format(location_filtered.iloc[0][cost_select])
            )

    procedure_data_df = pd.DataFrame(data=procedure_dict)

    return dash_table.DataTable(
        id="procedure-stats-table",
        columns=[{"name": i, "id": i} for i in procedure_dict.keys()],
        data=procedure_data_df.to_dict("records"),
        filter_action="native",
        sort_action="native",
        style_cell={
            "textOverflow": "ellipsis",
            "background-color": "#242a3b",
            "color": "#7b7d8d",
        },
        sort_mode="multi",
        page_size=5,
        style_as_list_view=False,
        style_header={"background-color": "#1f2536", "padding": "2px 12px 0px 12px"},
    )

@app.callback(
    Output("geo-map", "figure"),
    [
        Input("state-select", "value"),
        Input("metric-select", "value"),
        Input("year-slider", "value")
    ]
)
def update_choropleth_map(selected_states, selected_metrics, year_range):
    if not selected_states:
        selected_states = df_shark['State'].unique()
    
    # Filter by year range
    filtered_df = df_shark[
        (df_shark['Incident.year'] >= year_range[0]) & 
        (df_shark['Incident.year'] <= year_range[1])
    ]
    
    # Create incident counts and metrics
    metrics_data = {}
    for state in df_shark['State'].unique():
        state_data = filtered_df[filtered_df['State'] == state]
        metrics_data[state] = {
            'incident_count': len(state_data),
            'fatal_count': len(state_data[state_data['Victim.injury'] == 'fatal']),
            'avg_shark_length': state_data['Shark.length.m'].mean(),
            'avg_victim_age': state_data['Victim.age'].mean()
        }
    
    # Create DataFrame for choropleth
    plot_data = pd.DataFrame(metrics_data).T.reset_index()
    plot_data.columns = ['State'] + list(plot_data.columns[1:])
    
    # Dynamically generate state_code_mapping
    state_code_mapping = {state: str(idx + 1) for idx, state in enumerate(plot_data['State'].unique())}
    plot_data['state_code'] = plot_data['State'].map(state_code_mapping)
    
    # Create the choropleth
    fig = go.Figure(data=go.Choropleth(
        geojson=aus_states_geojson,
        locations=plot_data['state_code'],
        z=plot_data['incident_count'],
        locationmode='geojson-id',
        colorscale='Reds',
        colorbar_title="Incidents",
        text=[
            f"State: {row['State']}<br>" +
            f"Total Incidents: {row['incident_count']}<br>" +
            f"Fatal Incidents: {row['fatal_count']}<br>" +
            f"Avg Shark Length: {row['avg_shark_length']:.1f}m"
            for _, row in plot_data.iterrows()
        ],
        hoverinfo='text',
        marker_line_color='white',
        marker_line_width=0.5,
        featureidkey="properties.STATE_CODE"
    ))
    
    # Update layout
    fig.update_geos(
        visible=True,
        projection_type="mercator",
        center={"lat": -25.2744, "lon": 133.7751},
        lataxis_range=[-45, -10],
        lonaxis_range=[110, 155],
        bgcolor='#171b26',
        showcoastlines=True,
        coastlinecolor="White",
        showland=True,
        landcolor="lightgray",
        showocean=True,
        oceancolor="#171b26"
    )

    fig.update_layout(
        title_text=f"Shark Incidents ({year_range[0]}-{year_range[1]})",
        paper_bgcolor="#171b26",
        plot_bgcolor="#171b26",
        font={"color": "white"},
        height=600
    )

    return fig


if __name__ == "__main__":
    app.run_server(debug=False, host="0.0.0.0", port=8000)